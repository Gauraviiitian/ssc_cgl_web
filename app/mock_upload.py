from __future__ import annotations

import csv
import io

REQUIRED_FIELDS = ("question", "option_a", "option_b", "option_c", "option_d", "correct_option")
MAX_ROWS = 2000
MAX_FILE_BYTES = 2 * 1024 * 1024


class UploadTooLargeError(Exception):
    pass


def _normalize_header(name: str) -> str:
    return name.strip().lower().replace(" ", "_")


def _normalize_row(raw: dict, row_num: int) -> tuple[dict | None, str | None]:
    row = {_normalize_header(k): (v if v is not None else "") for k, v in raw.items()}
    row = {k: str(v).strip() for k, v in row.items()}

    missing = [f for f in REQUIRED_FIELDS if not row.get(f)]
    if missing:
        return None, f"row {row_num}: missing {', '.join(missing)}"

    correct = row["correct_option"].strip().upper()[:1]
    if correct not in ("A", "B", "C", "D"):
        return None, f"row {row_num}: correct_option must be A/B/C/D, got {row['correct_option']!r}"

    return {
        "text": row["question"],
        "option_a": row["option_a"],
        "option_b": row["option_b"],
        "option_c": row["option_c"],
        "option_d": row["option_d"],
        "correct_option": correct,
        "explanation": row.get("explanation", "").strip(),
    }, None


def _parse_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _parse_excel(file_bytes: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [_normalize_header(str(h)) for h in next(rows_iter)]
    rows = []
    for values in rows_iter:
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def parse_upload(file_bytes: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """Returns (valid_rows, errors). Never raises for bad row data — only for
    structurally unusable uploads (wrong extension, oversized file)."""
    if len(file_bytes) > MAX_FILE_BYTES:
        raise UploadTooLargeError(f"File is larger than {MAX_FILE_BYTES // (1024 * 1024)}MB")

    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext == "csv":
        raw_rows = _parse_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        raw_rows = _parse_excel(file_bytes)
    else:
        raise ValueError("Unsupported file type — upload a .csv, .xlsx, or .xls file")

    if len(raw_rows) > MAX_ROWS:
        raise UploadTooLargeError(f"File has more than {MAX_ROWS} rows")

    valid_rows: list[dict] = []
    errors: list[str] = []
    for i, raw in enumerate(raw_rows, start=2):  # row 1 is the header
        row, error = _normalize_row(raw, i)
        if error:
            errors.append(error)
        else:
            valid_rows.append(row)

    return valid_rows, errors
